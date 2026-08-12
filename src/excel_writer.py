"""Append-only writer for the approved BCG trend workbook."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Iterable

import yaml
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.cell_range import CellRange


class TemplateStructureError(RuntimeError):
    """Raised when editing would require guessing a workbook anchor."""


@dataclass(frozen=True)
class TrendItem:
    published_date: date
    source: str
    summary_ko: str
    note: str = ""
    url: str = ""

    @classmethod
    def from_mapping(cls, value: dict[str, Any]) -> "TrendItem":
        raw_date = value["published_date"]
        parsed = date.fromisoformat(raw_date) if isinstance(raw_date, str) else raw_date
        if not isinstance(parsed, date):
            raise TypeError("published_date must be an ISO date or date object")
        source = str(value.get("source", "")).strip()
        summary = str(value.get("summary_ko", "")).strip()
        if not source or not summary:
            raise ValueError("source and summary_ko are required")
        return cls(parsed, source, summary, str(value.get("note", "")), str(value.get("url", "")))


def load_rules(path: Path) -> dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8"))


def _rows_containing(ws, needle: str) -> list[int]:
    folded = needle.casefold()
    return sorted({
        cell.row for row in ws.iter_rows() for cell in row
        if cell.value is not None and folded in str(cell.value).casefold()
    })


def _single_row(ws, label: str) -> int:
    rows = _rows_containing(ws, label)
    if len(rows) != 1:
        raise TemplateStructureError(f"Expected one row containing {label!r}; found {rows}")
    return rows[0]


def _ordered_institution_rows(ws, markers: list[str], trend_row: int, impact_row: int) -> list[int]:
    """Find the last ordered marker sequence before the impact section."""
    selected: list[int] = []
    upper_bound = impact_row
    for marker in reversed(markers):
        candidates = [row for row in _rows_containing(ws, marker) if trend_row < row < upper_bound]
        if not candidates:
            raise TemplateStructureError(
                f"Ordered institution marker {marker!r} was not found before row {upper_bound}"
            )
        chosen = max(candidates)
        selected.append(chosen)
        upper_bound = chosen
    return list(reversed(selected))


def _locate_anchors(ws, rules: dict[str, Any]) -> tuple[int, int, int]:
    trend_row = _single_row(ws, rules["section_labels"]["trend"])
    impact_row = _single_row(ws, rules["section_labels"]["impact"])
    marker_rows = _ordered_institution_rows(ws, rules["institution_markers"], trend_row, impact_row)
    insertion_row = marker_rows[0]
    if not trend_row < insertion_row < impact_row:
        raise TemplateStructureError("Institution block must be between trend and impact sections")
    source_row = insertion_row - 1
    if source_row <= trend_row or all(ws.cell(source_row, col).value is None for col in range(1, 5)):
        raise TemplateStructureError("No normal trend row is available as a style source")
    return insertion_row, source_row, impact_row


def _shift_merges(ws, insertion_row: int, amount: int) -> list[CellRange]:
    source_merges: list[CellRange] = []
    original = [CellRange(str(item)) for item in ws.merged_cells.ranges]
    for merged in original:
        if merged.min_row == insertion_row - 1 and merged.max_row == insertion_row - 1:
            source_merges.append(merged)
        if merged.min_row >= insertion_row:
            ws.unmerge_cells(str(merged))
    for merged in original:
        if merged.min_row >= insertion_row:
            moved = CellRange(
                min_col=merged.min_col,
                min_row=merged.min_row + amount,
                max_col=merged.max_col,
                max_row=merged.max_row + amount,
            )
            ws.merge_cells(str(moved))
    return source_merges


def _copy_row_style(ws, source_row: int, target_row: int, max_column: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    ws.row_dimensions[target_row].hidden = ws.row_dimensions[source_row].hidden
    for col in range(1, max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if isinstance(source, MergedCell):
            continue
        target._style = copy(source._style)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def _shift_row_dimensions(ws, insertion_row: int, amount: int, max_row: int) -> None:
    """Shift row metadata that openpyxl's insert_rows intentionally ignores."""
    dimensions = {
        row: {
            "height": ws.row_dimensions[row].height,
            "hidden": ws.row_dimensions[row].hidden,
            "outline_level": ws.row_dimensions[row].outlineLevel,
            "collapsed": ws.row_dimensions[row].collapsed,
            "thick_top": ws.row_dimensions[row].thickTop,
            "thick_bottom": ws.row_dimensions[row].thickBot,
        }
        for row in range(insertion_row, max_row + 1)
    }
    for row in range(max_row, insertion_row - 1, -1):
        target = ws.row_dimensions[row + amount]
        source = dimensions[row]
        target.height = source["height"]
        target.hidden = source["hidden"]
        target.outlineLevel = source["outline_level"]
        target.collapsed = source["collapsed"]
        target.thickTop = source["thick_top"]
        target.thickBot = source["thick_bottom"]


def append_trends(
    template_path: Path,
    output_path: Path,
    items: Iterable[TrendItem | dict[str, Any]],
    rules: dict[str, Any],
    report_start: date | None = None,
    report_end: date | None = None,
    allow_empty: bool = False,
) -> Path:
    """Insert new trends while leaving historical cells untouched."""
    template_path = template_path.resolve()
    output_path = output_path.resolve()
    if template_path == output_path:
        raise ValueError("Output path must differ from the immutable template path")
    if template_path.suffix.lower() == ".xlsb":
        raise ValueError("Direct .xlsb editing is unsupported; use an approved .xlsx master")
    if template_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Template must be .xlsx or .xlsm")

    normalized = [item if isinstance(item, TrendItem) else TrendItem.from_mapping(item) for item in items]
    if not normalized and not allow_empty:
        raise ValueError("At least one new trend item is required")
    normalized.sort(key=lambda item: item.published_date)

    keep_vba = template_path.suffix.lower() == ".xlsm"
    wb = load_workbook(template_path, data_only=False, keep_vba=keep_vba, keep_links=True)
    sheet_name = rules["sheet_name"]
    if sheet_name not in wb.sheetnames:
        raise TemplateStructureError(f"Required sheet {sheet_name!r} was not found")
    ws = wb[sheet_name]
    insertion_row, source_row, _ = _locate_anchors(ws, rules)
    amount = len(normalized)
    source_merges: list[CellRange] = []
    if amount:
        original_max_row = ws.max_row
        source_merges = _shift_merges(ws, insertion_row, amount)
        _shift_row_dimensions(ws, insertion_row, amount, original_max_row)
        ws.insert_rows(insertion_row, amount)

    for offset, item in enumerate(normalized):
        row = insertion_row + offset
        _copy_row_style(ws, source_row, row, ws.max_column)
        ws.cell(row, 1, item.published_date)
        ws.cell(row, 2, item.source)
        ws.cell(row, 3, item.summary_ko)
        ws.cell(row, 4, item.note)
        if item.url:
            ws.cell(row, 3).hyperlink = item.url
        for merged in source_merges:
            translated = CellRange(
                min_col=merged.min_col,
                min_row=row,
                max_col=merged.max_col,
                max_row=row,
            )
            ws.merge_cells(str(translated))

    date_cell = rules.get("report_date_cell")
    if date_cell:
        if report_start is None or report_end is None:
            raise ValueError("report_start and report_end are required when report_date_cell is configured")
        ws[date_cell] = rules["report_date_format"].replace("%Y-%m-%d", "{}", 1).replace("%Y-%m-%d", "{}", 1).format(
            report_start.isoformat(), report_end.isoformat()
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    load_workbook(output_path, read_only=True, data_only=False, keep_vba=keep_vba).close()
    return output_path
