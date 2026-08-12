"""Inventory an Excel master template without modifying it."""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _matching_rows(ws, text: str) -> list[int]:
    needle = text.casefold()
    return sorted(
        {
            cell.row
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None and needle in str(cell.value).casefold()
        }
    )


def _ordered_institution_rows(
    ws, markers: list[str], trend_row: int | None, impact_row: int | None
) -> list[int] | None:
    if trend_row is None or impact_row is None:
        return None
    selected: list[int] = []
    upper_bound = impact_row
    for marker in reversed(markers):
        candidates = [row for row in _matching_rows(ws, marker) if trend_row < row < upper_bound]
        if not candidates:
            return None
        chosen = max(candidates)
        selected.append(chosen)
        upper_bound = chosen
    return list(reversed(selected))


def analyze_workbook(path: Path, rules: dict[str, Any]) -> dict[str, Any]:
    """Return a JSON-serializable structural report for *path*."""
    suffix = path.suffix.lower()
    if suffix == ".xlsb":
        raise ValueError(".xlsb is binary and cannot be safely round-tripped by openpyxl; convert to .xlsx")
    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError(f"Unsupported workbook type: {suffix}")

    wb = load_workbook(path, data_only=False, keep_vba=suffix == ".xlsm", keep_links=True)
    result: dict[str, Any] = {"path": str(path), "sheets": [], "has_vba_archive": wb.vba_archive is not None}
    for ws in wb.worksheets:
        style_counts = Counter(
            cell.style_id
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None or cell.has_style
        )
        merged = [str(item) for item in ws.merged_cells.ranges]
        sheet_report: dict[str, Any] = {
            "name": ws.title,
            "used_range": ws.calculate_dimension(),
            "merged_cells": merged,
            "column_widths": {
                get_column_letter(index): ws.column_dimensions[get_column_letter(index)].width
                for index in range(1, ws.max_column + 1)
            },
            "row_heights": {
                str(index): ws.row_dimensions[index].height
                for index in range(1, ws.max_row + 1)
                if ws.row_dimensions[index].height is not None
            },
            "style_distribution": {str(key): value for key, value in sorted(style_counts.items())},
            "drawing_count": len(ws._images) + len(ws._charts),
        }
        if ws.title == rules["sheet_name"]:
            trend_rows = _matching_rows(ws, rules["section_labels"]["trend"])
            impact_rows = _matching_rows(ws, rules["section_labels"]["impact"])
            institution_rows = {
                marker: _matching_rows(ws, marker) for marker in rules["institution_markers"]
            }
            trend_row = trend_rows[0] if len(trend_rows) == 1 else None
            impact_row = impact_rows[0] if len(impact_rows) == 1 else None
            ordered_rows = _ordered_institution_rows(
                ws, rules["institution_markers"], trend_row, impact_row
            )
            first_institution = ordered_rows[0] if ordered_rows else None
            data_start = trend_row + 2 if trend_row is not None else None
            data_end = first_institution - 1 if first_institution is not None else None
            sheet_report["sections"] = {
                "trend_label_rows": trend_rows,
                "impact_label_rows": impact_rows,
                "institution_rows": institution_rows,
                "selected_ordered_institution_rows": ordered_rows,
                "current_trend_data_range": (
                    f"A{data_start}:D{data_end}"
                    if data_start is not None and data_end is not None and data_end >= data_start
                    else None
                ),
                "institution_description_start_row": first_institution,
                "impact_section_row": impact_row,
            }
        result["sheets"].append(sheet_report)
    return result


def load_rules(path: Path) -> dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8"))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--rules", type=Path, default=Path("config/report_rules.yaml"))
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    report = analyze_workbook(args.workbook, load_rules(args.rules))
    payload = json.dumps(report, ensure_ascii=False, indent=2)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(payload + "\n", encoding="utf-8")
    print(payload)


if __name__ == "__main__":
    main()
