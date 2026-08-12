from __future__ import annotations

import json
from datetime import date, datetime, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.weekly_report import (
    WeeklyReportError,
    completed_week,
    generate_weekly_report,
    records_to_trends,
)


def _record(
    record_id: str = "bcg-ir-1",
    *,
    source: str = "BCG IR",
    title: str = "Official disclosure title",
    url: str = "https://example.invalid/disclosure/1",
) -> dict:
    return {
        "id": record_id,
        "published_date": "2026-08-06",
        "source": source,
        "title_original": title,
        "title_ko": None,
        "summary_ko": None,
        "document_number": "42/2026/CBTT-BCG",
        "url": url,
    }


def _payload(*, status: str = "SUCCESS", items: list[dict] | None = None) -> dict:
    return {
        "status": status,
        "requested_range": {"start": "2026-08-03", "end": "2026-08-09"},
        "sources": [
            {
                "source": "BCG IR",
                "status": status,
                "items": items or [],
                "errors": [],
            },
            {
                "source": "BCG Land IR",
                "status": status,
                "items": [],
                "errors": [],
            },
        ],
    }


def test_completed_week_is_previous_monday_through_sunday() -> None:
    assert completed_week(date(2026, 8, 10)) == (date(2026, 8, 3), date(2026, 8, 9))
    assert completed_week(date(2026, 8, 12)) == (date(2026, 8, 3), date(2026, 8, 9))


def test_records_use_source_title_link_and_exact_id_deduplication() -> None:
    item = _record()
    trends = records_to_trends(_payload(items=[item, dict(item)]))
    assert len(trends) == 1
    assert trends[0].summary_ko == "Official disclosure title"
    assert trends[0].note == "42/2026/CBTT-BCG"
    assert trends[0].url == "https://example.invalid/disclosure/1"


def test_excluded_exchange_source_is_rejected() -> None:
    payload = _payload()
    payload["sources"].append({"source": "HNX", "status": "SUCCESS", "items": [], "errors": []})
    with pytest.raises(WeeklyReportError, match="Excluded source"):
        records_to_trends(payload)


def test_generate_report_preserves_history_moves_sections_and_adds_link(
    master_workbook: Path,
    tmp_path: Path,
) -> None:
    before = load_workbook(master_workbook)
    before_ws = before.active
    historical = [[before_ws.cell(row, col).value for col in range(1, 5)] for row in (6, 7)]
    widths = {key: before_ws.column_dimensions[key].width for key in "ABCD"}
    source_styles = [before_ws.cell(7, col).style_id for col in range(1, 5)]
    source_borders = [str(before_ws.cell(7, col).border) for col in range(1, 5)]
    before.close()

    def collector(config: dict, **kwargs) -> dict:
        return _payload(items=[_record()])

    output_dir = tmp_path / "reports"
    metadata_dir = tmp_path / "history"
    result = generate_weekly_report(
        sources_path=Path("config/sources.yaml"),
        rules_path=Path("config/report_rules.yaml"),
        template_path=master_workbook,
        output_dir=output_dir,
        metadata_dir=metadata_dir,
        report_start=date(2026, 8, 3),
        report_end=date(2026, 8, 9),
        collector=collector,
        generated_at=datetime(2026, 8, 10, 0, 0, tzinfo=timezone.utc),
    )

    output = Path(result["output_path"])
    assert output.exists()
    assert result["item_count"] == 1
    assert Path(result["metadata_path"]).exists()
    assert json.loads(Path(result["metadata_path"]).read_text(encoding="utf-8"))["record_ids"] == ["bcg-ir-1"]

    wb = load_workbook(output)
    ws = wb.active
    assert [[ws.cell(row, col).value for col in range(1, 5)] for row in (6, 7)] == historical
    assert ws.cell(8, 1).value.date() == date(2026, 8, 6)
    assert ws.cell(8, 2).value == "BCG IR"
    assert ws.cell(8, 3).value == "Official disclosure title"
    assert ws.cell(8, 3).hyperlink.target == "https://example.invalid/disclosure/1"
    assert ws["A9"].value.startswith("HNX:")
    assert ws["A10"].value.startswith("HOSE:")
    assert ws["A11"].value.startswith("SSC:")
    assert ws["A13"].value == "2) 영향도 분석"
    assert {key: ws.column_dimensions[key].width for key in "ABCD"} == widths
    assert [ws.cell(8, col).style_id for col in range(1, 5)] == source_styles
    assert [str(ws.cell(8, col).border) for col in range(1, 5)] == source_borders
    assert all(ws.cell(8, col).alignment.wrap_text for col in range(1, 5))
    assert ws.row_dimensions[8].height == 42.0
    wb.close()

    reopened = load_workbook(output, read_only=True)
    reopened.close()


def test_successful_empty_week_creates_report_without_new_rows(
    master_workbook: Path,
    tmp_path: Path,
) -> None:
    def collector(config: dict, **kwargs) -> dict:
        return _payload(items=[])

    result = generate_weekly_report(
        sources_path=Path("config/sources.yaml"),
        rules_path=Path("config/report_rules.yaml"),
        template_path=master_workbook,
        output_dir=tmp_path / "reports",
        metadata_dir=tmp_path / "history",
        report_start=date(2026, 8, 3),
        report_end=date(2026, 8, 9),
        collector=collector,
    )
    wb = load_workbook(result["output_path"])
    assert wb.active.max_row == 14
    assert result["item_count"] == 0
    wb.close()


def test_failed_collection_writes_metadata_but_no_workbook(
    master_workbook: Path,
    tmp_path: Path,
) -> None:
    def collector(config: dict, **kwargs) -> dict:
        payload = _payload(status="FAILED")
        payload["sources"][0]["errors"] = [{"code": "network", "message": "unavailable"}]
        return payload

    output_dir = tmp_path / "reports"
    metadata_dir = tmp_path / "history"
    with pytest.raises(WeeklyReportError, match="All configured sources failed"):
        generate_weekly_report(
            sources_path=Path("config/sources.yaml"),
            rules_path=Path("config/report_rules.yaml"),
            template_path=master_workbook,
            output_dir=output_dir,
            metadata_dir=metadata_dir,
            report_start=date(2026, 8, 3),
            report_end=date(2026, 8, 9),
            collector=collector,
        )
    assert not list(output_dir.glob("*.xlsx"))
    assert (metadata_dir / "weekly-2026-08-09.json").exists()


def test_existing_weekly_output_is_never_overwritten(
    master_workbook: Path,
    tmp_path: Path,
) -> None:
    def collector(config: dict, **kwargs) -> dict:
        return _payload(items=[])

    kwargs = {
        "sources_path": Path("config/sources.yaml"),
        "rules_path": Path("config/report_rules.yaml"),
        "template_path": master_workbook,
        "output_dir": tmp_path / "reports",
        "metadata_dir": tmp_path / "history",
        "report_start": date(2026, 8, 3),
        "report_end": date(2026, 8, 9),
        "collector": collector,
        "generated_at": datetime(2026, 8, 10, 0, 0, tzinfo=timezone.utc),
    }
    first = generate_weekly_report(**kwargs)
    second = generate_weekly_report(**kwargs)
    assert first["output_path"] != second["output_path"]
    assert Path(first["output_path"]).exists()
    assert Path(second["output_path"]).exists()


def test_windows_scheduler_scripts_keep_the_approved_scope() -> None:
    runner = Path("scripts/run_weekly.ps1").read_text(encoding="utf-8")
    installer = Path("scripts/install_weekly_task.ps1").read_text(encoding="utf-8")
    assert '"-m", "src.weekly_report"' in runner
    assert '"--sources", "config\\sources.yaml"' in runner
    assert "New-ScheduledTaskTrigger -Weekly" in installer
    assert "-DaysOfWeek Monday" in installer
    assert "BCG Group Weekly Monitoring" in installer
    assert "HNX" not in runner + installer
    assert "HOSE" not in runner + installer
    assert "SSC" not in runner + installer
