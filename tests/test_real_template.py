from __future__ import annotations

import hashlib
import json
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from src.excel_template_analyzer import analyze_workbook
from src.excel_writer import append_trends, load_rules


ROOT = Path(__file__).resolve().parents[1]
MASTER = ROOT / "reports/template/BCG_Group_Trend_Master.xlsx"
RULES = ROOT / "config/report_rules.yaml"
DUMMY = ROOT / "tests/fixtures/dummy_items.json"


def _row_values(ws, row: int) -> list[object]:
    return [ws.cell(row, column).value for column in range(1, 5)]


def test_real_master_structure() -> None:
    report = analyze_workbook(MASTER, load_rules(RULES))
    sheet = report["sheets"][0]
    assert sheet["name"] == "BCG 동향"
    assert sheet["used_range"] == "A1:D55"
    assert sheet["merged_cells"] == []
    assert sheet["column_widths"] == {"A": 15.25, "B": 15.25, "C": 75.75, "D": 24.25}
    sections = sheet["sections"]
    assert sections["current_trend_data_range"] == "A5:D37"
    assert sections["selected_ordered_institution_rows"] == [38, 39, 40]
    assert sections["impact_section_row"] == 42


def test_real_master_writer_preserves_history_and_moves_metadata(tmp_path: Path) -> None:
    before_hash = hashlib.sha256(MASTER.read_bytes()).hexdigest()
    before = load_workbook(MASTER)
    before_ws = before["BCG 동향"]
    historical_values = {row: _row_values(before_ws, row) for row in range(3, 38)}
    historical_styles = {
        (row, column): before_ws.cell(row, column).style_id
        for row in range(3, 38)
        for column in range(1, 5)
    }
    downstream = {
        row: {
            "values": _row_values(before_ws, row),
            "styles": [before_ws.cell(row, column).style_id for column in range(1, 5)],
            "height": before_ws.row_dimensions[row].height,
            "hidden": before_ws.row_dimensions[row].hidden,
        }
        for row in range(38, 56)
    }
    widths = {column: before_ws.column_dimensions[column].width for column in "ABCD"}
    source_styles = [before_ws.cell(37, column).style_id for column in range(1, 5)]
    source_borders = [str(before_ws.cell(37, column).border) for column in range(1, 5)]
    before.close()

    output = tmp_path / "real-output.xlsx"
    items = json.loads(DUMMY.read_text(encoding="utf-8"))
    append_trends(
        MASTER,
        output,
        items,
        load_rules(RULES),
        report_start=date(2026, 8, 12),
        report_end=date(2026, 8, 19),
    )

    assert hashlib.sha256(MASTER.read_bytes()).hexdigest() == before_hash
    wb = load_workbook(output)
    ws = wb["BCG 동향"]
    for row in range(3, 38):
        assert _row_values(ws, row) == historical_values[row]
        assert [ws.cell(row, column).style_id for column in range(1, 5)] == [
            historical_styles[(row, column)] for column in range(1, 5)
        ]

    assert [ws.cell(row, 1).value.date() for row in range(38, 41)] == [
        date(2026, 8, 13), date(2026, 8, 14), date(2026, 8, 15)
    ]
    for row in range(38, 41):
        assert [ws.cell(row, column).style_id for column in range(1, 5)] == source_styles
        assert [str(ws.cell(row, column).border) for column in range(1, 5)] == source_borders
        assert ws.row_dimensions[row].height == 48.0

    for old_row, expected in downstream.items():
        new_row = old_row + 3
        assert _row_values(ws, new_row) == expected["values"]
        assert [ws.cell(new_row, column).style_id for column in range(1, 5)] == expected["styles"]
        assert ws.row_dimensions[new_row].height == expected["height"]
        assert ws.row_dimensions[new_row].hidden == expected["hidden"]

    assert ws["A41"].value.startswith("※ HNX")
    assert ws["A42"].value.startswith("※ HOSE")
    assert ws["A43"].value.startswith("※ SSC")
    assert ws["A45"].value == "2) 영향도 분석"
    assert ws["D2"].value == "2026-08-12 → 2026-08-19"
    assert {column: ws.column_dimensions[column].width for column in "ABCD"} == widths
    wb.close()

    reopened = load_workbook(output, read_only=True)
    assert reopened["BCG 동향"].max_row == 58
    reopened.close()
