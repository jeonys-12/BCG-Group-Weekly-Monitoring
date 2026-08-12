from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.excel_template_analyzer import analyze_workbook
from src.excel_writer import TemplateStructureError, append_trends


def _items() -> list[dict]:
    return [
        {"published_date": "2026-08-15", "source": "SSC", "summary_ko": "신규 3", "note": "SSSG 관련"},
        {"published_date": "2026-08-13", "source": "HOSE", "summary_ko": "신규 1", "note": ""},
        {"published_date": "2026-08-14", "source": "HNX", "summary_ko": "신규 2", "note": "유동성 관련"},
    ]


def test_analyzer_reports_required_structure(master_workbook: Path, rules: dict) -> None:
    report = analyze_workbook(master_workbook, rules)
    sheet = report["sheets"][0]
    assert sheet["name"] == "BCG 동향"
    assert sheet["used_range"] == "A1:D14"
    assert "A13:D14" in sheet["merged_cells"]
    assert sheet["column_widths"] == {"A": 14.0, "B": 18.0, "C": 64.0, "D": 28.0}
    assert sheet["row_heights"]["7"] == 42.0
    assert sheet["sections"]["current_trend_data_range"] == "A6:D7"
    assert sheet["sections"]["institution_description_start_row"] == 8
    assert sheet["sections"]["impact_section_row"] == 12
    assert sheet["style_distribution"]


def test_writer_preserves_history_and_moves_sections(master_workbook: Path, tmp_path: Path, rules: dict) -> None:
    before = load_workbook(master_workbook)
    before_ws = before["BCG 동향"]
    historical = [[before_ws.cell(row, col).value for col in range(1, 5)] for row in (6, 7)]
    widths = {key: before_ws.column_dimensions[key].width for key in "ABCD"}
    source_style_ids = [before_ws.cell(7, col).style_id for col in range(1, 5)]
    source_borders = [str(before_ws.cell(7, col).border) for col in range(1, 5)]
    before.close()

    output = tmp_path / "output.xlsx"
    append_trends(master_workbook, output, _items(), rules, report_start=date(2026, 8, 12), report_end=date(2026, 8, 19))
    wb = load_workbook(output)
    ws = wb["BCG 동향"]

    assert [[ws.cell(row, col).value for col in range(1, 5)] for row in (6, 7)] == historical
    assert [ws.cell(row, 1).value.date() for row in range(8, 11)] == [date(2026, 8, 13), date(2026, 8, 14), date(2026, 8, 15)]
    assert ws["A11"].value.startswith("HNX:")
    assert ws["A12"].value.startswith("HOSE:")
    assert ws["A13"].value.startswith("SSC:")
    assert ws["A15"].value == "2) 영향도 분석"
    assert ws["A16"].value == "기존 영향도 분석 문장"
    assert "A11:D11" in {str(value) for value in ws.merged_cells.ranges}
    assert "A15:D15" in {str(value) for value in ws.merged_cells.ranges}
    assert "A16:D17" in {str(value) for value in ws.merged_cells.ranges}
    assert {key: ws.column_dimensions[key].width for key in "ABCD"} == widths
    for row in range(8, 11):
        assert [ws.cell(row, col).style_id for col in range(1, 5)] == source_style_ids
        assert [str(ws.cell(row, col).border) for col in range(1, 5)] == source_borders
        assert all(ws.cell(row, col).alignment.wrap_text for col in range(1, 5))
        assert ws.row_dimensions[row].height == 42.0
    assert ws["D2"].value == "2026-08-12 → 2026-08-19"
    wb.close()

    reopened = load_workbook(output, read_only=True)
    assert reopened.sheetnames == ["BCG 동향"]
    reopened.close()


def test_template_is_never_overwritten(master_workbook: Path, rules: dict) -> None:
    with pytest.raises(ValueError, match="must differ"):
        append_trends(master_workbook, master_workbook, _items(), rules)


def test_missing_anchor_fails_instead_of_guessing(master_workbook: Path, tmp_path: Path, rules: dict) -> None:
    wb = load_workbook(master_workbook)
    wb["BCG 동향"]["A10"] = "기관 설명 누락"
    broken = tmp_path / "broken.xlsx"
    wb.save(broken)
    with pytest.raises(TemplateStructureError, match="institution marker"):
        append_trends(broken, tmp_path / "never.xlsx", _items(), rules)


def test_xlsb_is_rejected(tmp_path: Path, rules: dict) -> None:
    fake = tmp_path / "master.xlsb"
    fake.write_bytes(b"not an excel binary workbook")
    with pytest.raises(ValueError, match="xlsb"):
        append_trends(fake, tmp_path / "output.xlsx", _items(), rules)
