from __future__ import annotations

from copy import copy
from pathlib import Path

import pytest
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@pytest.fixture()
def rules() -> dict:
    return {
        "sheet_name": "BCG 동향",
        "section_labels": {"trend": "1) BCG 그룹 동향", "impact": "2) 영향도 분석"},
        "institution_markers": ["HNX", "HOSE", "SSC"],
        "report_date_cell": "D2",
        "report_date_format": "%Y-%m-%d → %Y-%m-%d",
    }


@pytest.fixture()
def master_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "master.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "BCG 동향"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    ws["A1"] = "BCG 그룹 동향 보고"
    ws["A1"].font = Font(name="맑은 고딕", size=16, bold=True)
    ws["A2"] = "보고 기준일"
    ws["D2"] = "2026-08-05 → 2026-08-12"
    ws.merge_cells("A4:D4")
    ws["A4"] = "1) BCG 그룹 동향"
    headers = ["날짜", "출처", "주요내용", "비고"]
    for col, value in enumerate(headers, 1):
        ws.cell(5, col, value)
    thin = Side(style="thin", color="222222")
    data_style = {
        "font": Font(name="맑은 고딕", size=10),
        "fill": PatternFill("solid", fgColor="FFF2CC"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "alignment": Alignment(vertical="top", wrap_text=True),
    }
    for row, values in {
        6: ["2026-08-01", "HOSE", "과거 자료 1", ""],
        7: ["2026-08-05", "SSC", "과거 자료 2", "기존 비고"],
    }.items():
        for col, value in enumerate(values, 1):
            cell = ws.cell(row, col, value)
            for key, style in data_style.items():
                setattr(cell, key, copy(style))
            if col == 1:
                cell.number_format = "yyyy-mm-dd"
        ws.row_dimensions[row].height = 42
    ws["A8"] = "HNX: Hanoi Stock Exchange 설명"
    ws.merge_cells("A8:D8")
    ws["A9"] = "HOSE: Ho Chi Minh Stock Exchange 설명"
    ws.merge_cells("A9:D9")
    ws["A10"] = "SSC: State Securities Commission 설명"
    ws.merge_cells("A10:D10")
    ws.merge_cells("A12:D12")
    ws["A12"] = "2) 영향도 분석"
    ws.merge_cells("A13:D14")
    ws["A13"] = "기존 영향도 분석 문장"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 64
    ws.column_dimensions["D"].width = 28
    ws.row_dimensions[8].height = 24
    ws.row_dimensions[12].height = 30
    wb.save(path)
    return path

@pytest.fixture(autouse=True)
def block_unmarked_live_network(request, monkeypatch) -> None:
    """Guarantee that default unit/regression tests make zero network calls."""

    if request.node.get_closest_marker("integration") is not None:
        return

    def blocked(*args, **kwargs):
        raise AssertionError("Live network access is forbidden outside integration tests")

    monkeypatch.setattr(requests.sessions.Session, "request", blocked)