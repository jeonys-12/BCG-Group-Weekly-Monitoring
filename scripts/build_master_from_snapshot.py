"""Build a DRM-free .xlsx master from an Excel COM formatting snapshot."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter


HORIZONTAL = {-4131: "left", -4108: "center", -4152: "right", -4130: "justify", 7: "centerContinuous"}
VERTICAL = {-4160: "top", -4108: "center", -4107: "bottom", -4130: "justify"}
UNDERLINE = {-4142: None, 2: "single", -4119: "double", 5: "singleAccounting", 4: "doubleAccounting"}
LINE_STYLE = {-4142: None, 1: "thin", -4115: "dashed", 4: "dashDot", 5: "dashDotDot", -4118: "dotted", -4119: "double", 13: "slantDashDot"}
WEIGHT_STYLE = {1: "hair", 2: "thin", -4138: "medium", 4: "thick"}


def _side(value: dict[str, Any]) -> Side:
    style = LINE_STYLE.get(value["line_style"])
    if style == "thin" or style is None:
        style = WEIGHT_STYLE.get(value["weight"], style)
    return Side(style=style, color=value.get("color") if style else None)


def build_master(snapshot_path: Path, output_path: Path) -> Path:
    snapshot = json.loads(snapshot_path.read_text(encoding="utf-8"))
    if snapshot["has_vba"] or snapshot["shapes"] or snapshot["hyperlinks"] or snapshot["conditional_formats"]:
        raise ValueError("Snapshot contains unsupported features; manual review is required")

    wb = Workbook()
    ws = wb.active
    ws.title = snapshot["sheet_name"]
    ws.sheet_view.showGridLines = snapshot["show_gridlines"]
    ws.freeze_panes = None

    for item in snapshot["columns"]:
        dimension = ws.column_dimensions[get_column_letter(item["column"])]
        dimension.width = item["width"]
        dimension.hidden = item["hidden"]
    for item in snapshot["rows"]:
        dimension = ws.row_dimensions[item["row"]]
        dimension.height = item["height"]
        dimension.hidden = item["hidden"]

    for item in snapshot["cells"]:
        cell = ws.cell(item["row"], item["column"])
        cell.value = item["formula"] or item["value"]
        cell.number_format = item["number_format"] or "General"
        font = item["font"]
        cell.font = Font(
            name=font["name"], size=font["size"], bold=font["bold"], italic=font["italic"],
            underline=UNDERLINE.get(font["underline"]), strike=font["strike"], color=font.get("color"),
        )
        fill = item["fill"]
        cell.fill = PatternFill(
            fill_type="solid" if fill["pattern"] == 1 else None,
            fgColor=fill.get("color") or "000000",
            bgColor=fill.get("pattern_color") or "000000",
        )
        alignment = item["alignment"]
        rotation = alignment["text_rotation"] if 0 <= alignment["text_rotation"] <= 180 else 0
        cell.alignment = Alignment(
            horizontal=HORIZONTAL.get(alignment["horizontal"]),
            vertical=VERTICAL.get(alignment["vertical"]),
            wrap_text=alignment["wrap_text"], shrink_to_fit=alignment["shrink_to_fit"],
            indent=alignment["indent"], text_rotation=rotation,
        )
        borders = item["borders"]
        cell.border = Border(
            left=_side(borders["7"]), top=_side(borders["8"]),
            bottom=_side(borders["9"]), right=_side(borders["10"]),
        )
        cell.protection = Protection(
            locked=item["protection"]["locked"], hidden=item["protection"]["hidden"]
        )

    for merged in snapshot["merges"]:
        ws.merge_cells(merged)

    page = snapshot["page_setup"]
    ws.page_setup.orientation = "portrait" if page["orientation"] == 1 else "landscape"
    ws.page_setup.paperSize = str(page["paper_size"])
    ws.page_setup.fitToWidth = int(page["fit_to_pages_wide"]) if isinstance(page["fit_to_pages_wide"], int) else 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = page["left_margin"] / 72
    ws.page_margins.right = page["right_margin"] / 72
    ws.page_margins.top = page["top_margin"] / 72
    ws.page_margins.bottom = page["bottom_margin"] / 72
    ws.page_margins.header = page["header_margin"] / 72
    ws.page_margins.footer = page["footer_margin"] / 72
    ws.print_options.horizontalCentered = page["center_horizontally"]
    ws.print_options.verticalCentered = page["center_vertically"]
    if page["print_area"]:
        ws.print_area = page["print_area"].replace("$", "")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("snapshot", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    print(build_master(args.snapshot, args.output))


if __name__ == "__main__":
    main()
